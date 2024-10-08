import csv
import openpyxl
from datetime import datetime


# Вираховування віку
def calculate_age(birth_date_str):
    birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d")
    today = datetime.today()
    return today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))


# Зчитування .csv файлу з використанням list comprehension
def read_csv_file(file_path):
    try:
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            return [
                {
                    'surname': row['Прізвище'],
                    'name': row['Ім’я'],
                    'middle_name': row['По батькові'],
                    'gender': row['Стать'],
                    'birth_date': row['Дата народження'],
                    'age': calculate_age(row['Дата народження'])
                }
                for row in csv.DictReader(csvfile)
            ]
    except FileNotFoundError:
        print("Файл не знайдено")
        exit()
    except csv.Error:
        print("Помилка зчитування .csv")
        return None
    except Exception as e:
        print(f"Помилка: {e}")
        return None


# Створення файлу .xlsx
def create_xlsx_file(users, file_path):
    try:
        workbook = openpyxl.Workbook()
        sheets = {
            "all": workbook.active,
            "younger_18": workbook.create_sheet("younger_18"),
            "18-45": workbook.create_sheet("18-45"),
            "45-70": workbook.create_sheet("45-70"),
            "older_70": workbook.create_sheet("older_70")
        }
        sheets["all"].title = "all"
        headers = ['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік']

        # Додаємо заголовки на всі аркуші
        [sheet.append(headers) for sheet in sheets.values()]

        # Додаємо дані користувачів та розподіляємо по аркушах
        [sheets["younger_18" if user['age'] < 18 else
                "18-45" if 18 <= user['age'] <= 45 else
                "45-70" if 46 <= user['age'] <= 70 else
                "older_70"].append([index, user['surname'], user['name'],
                                   user['middle_name'], user['birth_date'], user['age']])
         for index, user in enumerate(users, start=1)]

        workbook.save(file_path)
    except Exception as e:
        print(f"Помилка: {e}")


def main():
    csv_file_path = 'employees.csv'
    xlsx_file_path = 'employees_data.xlsx'
    users = read_csv_file(csv_file_path)
    # Перевірка на присутність підходящих даних
    if users:
        create_xlsx_file(users, xlsx_file_path)
        print("OK")
    else:
        print("Підходящі дані відсутні")


if __name__ == '__main__':
    main()
